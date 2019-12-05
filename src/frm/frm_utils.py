
import datetime
import math
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, Series
import pulp
import pandas as pd
from pandas.api.types import is_datetime64_any_dtype as is_datetime
import plotly.graph_objects as go
from fuzzywuzzy import process # library for fuzzy string matching

def find_keyword(sheet: Worksheet, keyword: str, head_col: int) -> (int, int):
    "return coordinates for a certain word (e.g., origin or destination) in an excel sheet"
    if sheet is not None:
        for row in range(1, sheet.max_row+1):
            for column in range(head_col, sheet.max_column+1):
                cell = sheet.cell(row=row, column=column)
                if keyword.lower() in str(cell.value).lower():
                    return (cell.row, cell.column - 1) # Excel starts with 1, python from 0
    else:
        raise ValueError("This worksheet doesn't exist")

def _clean_col(df: DataFrame) -> None:
    "clean illegal chars in column names"
    try:
        column_names = []
        for name in list(df.columns):
            for c in r"(\&":
                if c in name: name = name[:name.find(c)]
            column_names.append(name.strip().lower())
        df.columns = column_names
    except AttributeError:
        pass

def _convert_datetime(df: DataFrame) -> None:
    "convert the time variables to the datetime format based on ops day"
    try:
        today = pd.Timestamp('today')
        time_var = process.extractOne('time', df.columns)[0]
        day_var = process.extractOne('ops day', df.columns)[0]
        day_sequence = sorted(df[day_var].unique())

        if not is_datetime(df.loc[:, time_var]): # only do this when we got datetime.time input

            df.loc[:, time_var] = pd.to_datetime(df[time_var].astype(str)).map(lambda x: x.replace(year=today.year, month=today.month, day=today.day))

            if len(day_sequence) == 2:
                df.loc[df[day_var] == day_sequence[1], time_var] = df[time_var] + pd.Timedelta(days=1)
            
            df.loc[:, time_var] = df[time_var].astype('datetime64[s]')

    except ValueError:
        pass

def forecast(df: DataFrame, growth_rate: float, num_years: int) -> DataFrame:
    "apply the growth rate on the current volumes for a number of years in the future"
    volume_var = process.extractOne('volume', df.columns)[0]
    df.loc[:, volume_var] = df[volume_var] * ((1 + growth_rate) ** num_years)
    return df

def replace_colname(df: DataFrame, new_words: list) -> DataFrame:
    "standardize column names for later merge and concatenations"
    if df is not None:
        for word in new_words:
            replacement = {process.extractOne(word, df.columns)[0]:word}
            df.rename(columns = replacement, inplace=True)
        return df[new_words]
    else:
        return df    

def ws_to_df(sheet: Worksheet, header_row:int, start_col:int, end_col:int) -> DataFrame:
    "Select data from a worksheet and return a DataFrame"
    data = sheet.values
    cols = [next(data) for k in range(header_row)][-1] # skip non-headers
    COL = cols[start_col:end_col+1]

    DATA = list(data)
    data = [row[start_col:end_col+1] for row in DATA]
    df = DataFrame(data, columns = COL)
    df.dropna(subset=[COL[0]], inplace=True)
    if 'schedule' in sheet.title.lower():
        _clean_col(df)
        _convert_datetime(df)
    return df

def hourly_volume_graph_in_out(df_inbound: DataFrame, df_outbound: DataFrame, preparedtime, schedule_type: str):
    "Plot hourly in/out volumes for different transportation types, i.e., linehaul, pud, and air"
    volume_arr = process.extractOne('volume', df_inbound.columns)[0] #fuzzy string matching
    time_arr = process.extractOne('arrival time', df_inbound.columns)[0]
    df_inbound.loc[:, 'offload time'] = df_inbound[time_arr] + pd.Timedelta(minutes=preparedtime)
    involume_hour = df_inbound[[volume_arr, 'offload time']].resample('1H', on='offload time').sum() 

    volume_dep = process.extractOne('volume', df_outbound.columns)[0]
    time_dep = process.extractOne('departure time', df_outbound.columns)[0]
    df_outbound.loc[:, 'onload time'] = df_outbound[time_dep] - pd.Timedelta(minutes=preparedtime)
    outvolume_hour = df_outbound[[volume_dep, 'onload time']].resample('1H', on='onload time').sum()

    volume_hour = pd.concat([involume_hour, outvolume_hour], axis=1, sort=False).fillna(0)
    volume_hour.columns = ['in', 'out']
    #volume_hour.index = volume_hour.index.time
    volume_hour.index.name = 'time'
    barchart = go.Figure(data=[
        go.Bar(name='inbound', x=volume_hour.index, y=volume_hour['in']),
        go.Bar(name='outbound', x=volume_hour.index, y=volume_hour['out'])
    ])
    barchart.update_layout(barmode='group', title_text=f'Hourly volume graph for {schedule_type} schedule(s)')
    #barchart = volume_hour.iplot(kind='bar', color=['b', 'r'], grid=True, title=f'Hourly volume graph for {schedule_type} schedule(s)')

    return barchart

def hourly_table_arr_dep(df_inbound: DataFrame, df_outbound: DataFrame):
    "Create the table to see simultaneously arrival/departure per hour movements"
    time_arr = process.extractOne('arrival time', df_inbound.columns)[0]
    origin = process.extractOne('origin', df_inbound.columns)[0]
    origin_hourly = df_inbound.groupby(origin).resample('1H', on=time_arr).size()
    origin_hourly = origin_hourly[origin_hourly!=0]
    origin_hourly.sort_values(ascending=False, inplace=True)

    time_dep = process.extractOne('departure time', df_outbound.columns)[0]
    dest = process.extractOne('destination', df_outbound.columns)[0]
    dest_hourly = df_outbound.groupby(dest).resample('1H', on=time_dep).size()
    dest_hourly = dest_hourly[dest_hourly!=0]
    dest_hourly.sort_values(ascending=False, inplace=True)

    return origin_hourly, dest_hourly

def solve_lp_model(volume_2bsort: Series, max_buffer = None, min_capacity = None, max_capacity = None) -> DataFrame:
    "Calculate the linear programming problem for the sorter: find the minimum sorter capacity"
    results = {}
    sort_model = pulp.LpProblem("sort capacity optimization", sense=pulp.LpMinimize)
    x = pulp.LpVariable.dicts("x", volume_2bsort.index, cat='Integer')
    capacity = pulp.LpVariable("capacity", min_capacity, max_capacity, cat='Integer')

    #sort_model += pulp.lpSum(x[i] for i in volume_2bsort.nlargest(4).index) + x[volume_2bsort.index[-1]], "Minimize_sort_capacity"
    sort_model += capacity, "Minimize_sort_capacity"

    for i in volume_2bsort.index:
        sort_model += x[i] >= 0
        sort_model += x[i] <= capacity
        sort_model += x[i] <= volume_2bsort[volume_2bsort.index[0:(volume_2bsort.index.get_loc(i)+1)]].sum() \
            - pulp.lpSum(x[j] for j in volume_2bsort.index[0:volume_2bsort.index.get_loc(i)])
        if max_buffer is not None:
            sort_model += volume_2bsort[volume_2bsort.index[0:(volume_2bsort.index.get_loc(i)+1)]].sum() \
                - pulp.lpSum(x[j] for j in volume_2bsort.index[0:volume_2bsort.index.get_loc(i)+1]) <= max_buffer

    sort_model += pulp.lpSum(x[i] for i in volume_2bsort.index) == volume_2bsort.sum(), "Total volume to be sorted"

    sort_model.solve()
    print(f"Model status: {pulp.LpStatus[sort_model.status]}")
    print(f"Optimal capacity: {int(pulp.value(sort_model.objective))}")
    for v in sort_model.variables():
        results[v.name] = v.varValue
        #print(str(v.name) + " = " + str(v.varValue))
    results = pd.DataFrame.from_dict(results, orient='index').iloc[1:, :]
    results.index = pd.to_datetime(results.index.str.replace('_', '-').map(lambda x: x[1:]))
    results.sort_index(inplace=True)
    results.columns = ('sorted',)
    results.loc[:, 'sta'] = volume_2bsort
    for i in results.index:
        if i == results.index[0]:
            results.loc[i, 'buffer'] = volume_2bsort[i] - results.loc[i, 'sorted']
        else:
            results.loc[i, 'buffer'] = volume_2bsort[volume_2bsort.index[0:(volume_2bsort.index.get_loc(i)+1)]].sum() - \
                results.loc[results.index[0:(results.index.get_loc(i)+1)], 'sorted'].sum()
    return results.loc[:, ['sta', 'sorted', 'buffer']].astype(int)

def calc_volume_spread(volume_sorted, volume_availability, dest_distribution, start_time, end_time, method, spread_func, ts):
    "calculate destination spread of volumes for each sort period"
    time_index = [t for t in volume_availability.columns if type(t) is pd.Timestamp]
    volume_ts = volume_availability.loc[:, ['level_0'] + ['origin'] + ['arrival time'] + ['id'] + [
        t for t in time_index if t > start_time and t <= end_time] + (
            ['departure time'] if 'departure time' in volume_availability.columns else [])].copy()
    volume_ts.loc[:, 'sum'] = volume_ts.iloc[:, 4:-1 if 'departure time' in volume_availability.columns else None].sum(axis=1)
    volume_ts = volume_ts[volume_ts['sum'] > 0]
    if method == 'earliest_dep_first':
        volume_ts.sort_values('departure time', inplace=True)
    elif method == 'random':
        volume_ts = volume_ts.sample(frac=1) # randomize the movements
    elif method == 'fifo':
        volume_ts.sort_values('arrival time', inplace=True)
    mov_selected = volume_ts[volume_ts['sum'].cumsum() <= volume_sorted[end_time]]
    mov_selected = mov_selected.merge(mov_selected['sum'].cumsum(), left_index=True, right_index=True, suffixes=(
        'total', 'cum'), sort=False)
    leftover = volume_sorted[end_time] - mov_selected['sumtotal'].sum()
    if (volume_ts['sum'].sum() - volume_sorted[end_time]) > 0: # if the available volumes cannot all be sorted
        mov_left = volume_ts[volume_ts['sum'].cumsum() > volume_sorted[end_time]]
        mov_left = mov_left.merge(mov_left['sum'].cumsum(), left_index=True, right_index=True, suffixes=(
            'total', 'cum'), sort=False)
        for index, mov in mov_left.iterrows():
            if mov['sumcum'] == mov_left.iloc[mov_left.sumcum.searchsorted(leftover, side='right'), -1]:
                volume_availability.loc[index, end_time + pd.Timedelta(ts)] += (mov['sumcum'] - leftover)
                mov['sumtotal'] = leftover
                mov_selected = mov_selected.append(mov)
            else:
                volume_availability.loc[index, end_time + pd.Timedelta(ts)] += mov['sumtotal']
    volume_availability.loc[:, [
        t for t in time_index if t > start_time and t <= end_time]] = 0 # clean up sorted volumes
    inbound_dist = (mov_selected[['level_0', 'sumtotal']].groupby('level_0').sum()/mov_selected['sumtotal'].sum()).T
    inbound_dist.index = (end_time, )
    spread = mov_selected.apply(spread_func, axis=1)
    dest_spread = pd.DataFrame(spread.sum(axis=0)).T
    dest_spread.index = (end_time, )
    mov_linehaul = mov_selected[mov_selected['level_0'] == 'linehaul']
    inbound_linehaul_volume = mov_linehaul['sumtotal'].sum()
    if inbound_linehaul_volume != 0:
        linehaul_spread = mov_linehaul.apply(spread_func, axis=1)
        linehaul_to_air = linehaul_spread[[col for col in linehaul_spread.columns if 'air' in col]].sum().sum()
        xray = linehaul_to_air/inbound_linehaul_volume
    else:
        xray = 0
    xray_tput = {end_time: xray}

    return volume_availability, dest_spread, inbound_dist, xray_tput