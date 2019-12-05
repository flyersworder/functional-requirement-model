"Skeleton of the FRM model"
import datetime
import math
import yaml
import pandas as pd
import numpy as np
from pathlib import Path
from fuzzywuzzy import process
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from frm.frm_utils import find_keyword, ws_to_df, hourly_volume_graph_in_out, hourly_table_arr_dep, replace_colname, solve_lp_model, forecast, calc_volume_spread

class FRM():
    "Skeleton class"

    def __init__(self, excel_file, config_file=None):

        file_path = Path(__file__)
        config_path = file_path.parents[0] / Path('config.yaml')
        with config_path.open('r') as fobj:
            _config = yaml.safe_load(fobj)

        self.PREP_TIME = _config['PREP_TIME']
        self.GROWTH_RATE = _config['GROWTH_RATE']
        self.MAX_BUFFER = _config['MAX_BUFFER']
        self.OUTBOUND_BUFFER = _config['OUTBOUND_BUFFER']
        self.PPH_OFFLOAD = _config['PPH_OFFLOAD']
        self.PPH_LOAD = _config['PPH_LOAD']
        self.PIECE_PER_BAG = _config['UNIT_CONVERTION']['PIECE_PER_BAG']
        self.PARCEL_PER_CAGE_DG = _config['UNIT_CONVERTION']['PARCEL_PER_CAGE_DG']
        self.PARCEL_PER_CAGE_NCY = _config['UNIT_CONVERTION']['PARCEL_PER_CAGE_NCY']
        self.HPC = _config['UNIT_CONVERTION']['HPC']

        # load excel file
        print(f"Start loading excel file from {excel_file}")
        self.work_book = load_workbook(excel_file, data_only=True)
        schedule_sheets = filter(
            lambda x: 'schedule' in x.lower(), self.work_book.sheetnames)

        self.lh_schedule_sheet = None
        self.pud_schedule_sheet = None
        self.air_schedule_sheet = None
        self.orig_dest_sheet = None
        self.vol_dist_sheet = None # movement data for orig-dest relationship if available

        # load schedule sheets
        for sheet in schedule_sheets:
            if 'lh' in sheet.lower():
                print(f'load schedule sheet: {sheet}')
                self.lh_schedule_sheet = self.work_book[sheet]
            elif 'pud' in sheet.lower():
                print(f'load schedule sheet: {sheet}')
                self.pud_schedule_sheet = self.work_book[sheet]
            elif 'air' in sheet.lower():
                print(f'load schedule sheet: {sheet}')
                self.air_schedule_sheet = self.work_book[sheet]
            else:
                raise ValueError(f"unknown schedule sheet type {sheet}")

        # load origin & destination relationship
        for sheet in self.work_book.sheetnames:
            if 'orig' in sheet.lower() or 'dest' in sheet.lower():
                print(f'load relationship sheet: {sheet}')
                self.orig_dest_sheet = self.work_book[sheet]
            elif 'dist' in sheet.lower():
                print(f'load volume distribution sheet: {sheet}')
                self.vol_dist_sheet = self.work_book[sheet]

        if not self.orig_dest_sheet and not self.vol_dist_sheet: print('No origin & destination relationship with the schedule')

        print(f"Done loading excel file")

        self.schedule_sheets = {
            'linehaul': self.lh_schedule_sheet,
            'pud': self.pud_schedule_sheet,
            'air': self.air_schedule_sheet,
        }

    counter_inbound = 0 # make sure it only prints once
    def get_inbound_df(self, schedule_type='linehaul', num_years=0):
        "Get pandas dataframe out of inbound schedules"
        FRM.counter_inbound += 1
        schedule_sheet = self.schedule_sheets.get(schedule_type)

        if schedule_sheet:
            header_row, start_col = find_keyword(schedule_sheet, 'origin', 1)
            end_col = find_keyword(schedule_sheet, 'sum', 1)[1]
            inbound_df = ws_to_df(schedule_sheet, header_row, start_col, end_col).fillna(0).replace('NA', 0)
            if num_years > 0:
                if FRM.counter_inbound==1: print(f'Forecasting for {num_years} years with growth rate {self.GROWTH_RATE}')
                inbound_df = forecast(inbound_df, self.GROWTH_RATE, num_years)
        else:
            inbound_df = None

        return inbound_df

    counter_outbound = 0
    def get_outbound_df(self, schedule_type='linehaul', num_years=0):
        "Get pandas dataframe out of outbound schedules"
        FRM.counter_outbound += 1
        schedule_sheet = self.schedule_sheets.get(schedule_type)

        if schedule_sheet:
            head_col = find_keyword(schedule_sheet, 'sum', 1)[1] + 2
            header_row, start_col = find_keyword(schedule_sheet, 'destination', head_col)
            end_col = find_keyword(schedule_sheet, 'sum', head_col)[1]
            outbound_df = ws_to_df(schedule_sheet, header_row, start_col, end_col).fillna(0).replace('NA', 0)
            if num_years > 0:
                if FRM.counter_outbound==1: print(f'Forecasting for {num_years} years with growth rate {self.GROWTH_RATE}')
                outbound_df = forecast(outbound_df, self.GROWTH_RATE, num_years)
        else:
            outbound_df = None

        return outbound_df

    def get_orig_dest_df(self):
        "Get pandas dataframe out of the origin & destination relationship"
        header_row, start_col = find_keyword(self.orig_dest_sheet, 'origin', 1)
        end_col = self.orig_dest_sheet.max_column

        distribution = ws_to_df(self.orig_dest_sheet, header_row, start_col, end_col)
        distribution.columns = distribution.columns.str.lower()

        dest_df = self.combine_outbound_df()
        dest_df = dest_df.applymap(lambda x: x.lower().strip() if type(x) == str else x)
        dest_df.reset_index(inplace=True)
        
        for name in list(distribution.columns)[1:-1]:
            new_name = str(dest_df.loc[dest_df['destination']==name, 'level_0'].values[0]) + '-' + str(
                dest_df.loc[dest_df['destination']==name, 'load type'].values[0]) + '-' + name
            replacement = {name: new_name}
            distribution.rename(columns = replacement, inplace=True)

        distribution = distribution.iloc[:, :-1]
        return distribution.drop_duplicates().reset_index(drop=True)       

    def get_vol_dist_df(self):
        "Get pandas dataframe out of the volume distribution data based on movements"
        header_row, start_col = find_keyword(self.vol_dist_sheet, 'from', 1)
        end_col = find_keyword(self.vol_dist_sheet, 'deptime', 1)[1]

        return ws_to_df(self.vol_dist_sheet, header_row, start_col, end_col)

    def hourly_volume_graph(self, schedule_type='linehaul', break_point=None, num_years=0, display=False):
        "Plot the hourly in/out volume graph for a specific schedule type, i.e., linehaul, pud, or air, or all schedules combined"
        if schedule_type == 'all':
            df_inbound = self.combine_inbound_df(break_point, num_years).iloc[:, :-1]
            df_outbound = self.combine_outbound_df(break_point, num_years).iloc[:, :-1]
        else:
            df_inbound = self.get_inbound_df(schedule_type, num_years)
            df_outbound = self.get_outbound_df(schedule_type, num_years)

        barchart = hourly_volume_graph_in_out(df_inbound, df_outbound, self.PREP_TIME, schedule_type)
        
        if display is True:
            barchart.show()

        return barchart

    def hourly_arr_dep_table(self, schedule_type='linehaul', num_years=0):
        "Create the hourly arr/dep table per orig/dest for a specific schedule type, i.e., linehaul, pud, or air"
        df_inbound = self.get_inbound_df(schedule_type, num_years)
        df_outbound = self.get_outbound_df(schedule_type, num_years)
        origin_hourly, dest_hourly = hourly_table_arr_dep(df_inbound, df_outbound)
        return origin_hourly, dest_hourly

    def combine_inbound_df(self, break_point=None, num_years=0):
        "Combine all the inbound dfs into one df"
        colnames = ['origin', 'arrival time', 'vehicle type', 'load type', 'volume', \
            'documents', 'box', 'smalls', 'dangerous goods', 'ncoy', 'hpc', 'freight', 'sum']
        
        lh_inbound_df = replace_colname(self.get_inbound_df(schedule_type='linehaul', num_years=num_years), colnames)
        pud_inbound_df = replace_colname(self.get_inbound_df(schedule_type='pud', num_years=num_years), colnames)
        air_inbound_df = replace_colname(self.get_inbound_df(schedule_type='air', num_years=num_years), colnames)
        cinbound_df = pd.concat([lh_inbound_df, pud_inbound_df, air_inbound_df], keys=['linehaul', 'pud', 'air'])
        # if sum equals to 1 it has no effect, otherwise it converts volumes to percentages
        cinbound_df.iloc[:, 5:] = cinbound_df.iloc[:, 5:].apply(lambda x: x/x['sum'], axis=1)

        if break_point is not None:
            break_point = pd.to_datetime(break_point)
            cinbound_df.loc[cinbound_df['arrival time'] < break_point, 'arrival time'] += pd.Timedelta(days=1)

        cinbound_df.loc[:, 'offload time'] = cinbound_df['arrival time'] + pd.Timedelta(minutes=self.PREP_TIME)
        
        return cinbound_df

    def combine_outbound_df(self, break_point=None, num_years=0):
        "Combine all the outbound dfs into one df"
        colnames = ['destination', 'departure time', 'vehicle type', 'load type', 'volume', \
            'documents', 'box', 'smalls', 'dangerous goods', 'ncoy', 'hpc', 'freight', 'sum']
        
        lh_outbound_df = replace_colname(self.get_outbound_df(schedule_type='linehaul', num_years=num_years), colnames)
        pud_outbound_df = replace_colname(self.get_outbound_df(schedule_type='pud', num_years=num_years), colnames)
        air_outbound_df = replace_colname(self.get_outbound_df(schedule_type='air', num_years=num_years), colnames)
        coutbound_df = pd.concat([lh_outbound_df, pud_outbound_df, air_outbound_df], keys=['linehaul', 'pud', 'air'])
        # if sum equals to 1 it has no effect, otherwise it converts volumes to percentages
        coutbound_df.iloc[:, 5:] = coutbound_df.iloc[:, 5:].apply(lambda x: x/x['sum'], axis=1)

        if break_point is not None:
            break_point = pd.to_datetime(break_point)
            coutbound_df.loc[coutbound_df['departure time'] < break_point, 'departure time'] += pd.Timedelta(days=1)

        coutbound_df.loc[:, 'finished load time'] = coutbound_df['departure time'] - pd.Timedelta(minutes=self.PREP_TIME)

        return coutbound_df

    def descriptive_stats(self, break_point=None, num_years=0):
        "Some descriptive stats for inbound and outbound volumes per material type"
        combine_inbound_df = self.combine_inbound_df(break_point, num_years)
        combine_outbound_df = self.combine_outbound_df(break_point, num_years)
        inbound_stats = {}
        outbound_stats = {}
        inbound_stats['total volume'] = combine_inbound_df['volume'].sum()
        outbound_stats['total volume'] = combine_outbound_df['volume'].sum()
        for mat in ['documents', 'box', 'smalls', 'dangerous goods', 'ncoy', 'hpc', 'freight']:
            inbound_stats[mat] = (combine_inbound_df['volume'] * combine_inbound_df[mat]).sum()
            outbound_stats[mat] = (combine_outbound_df['volume'] * combine_outbound_df[mat]).sum()
        
        inbound = pd.DataFrame.from_dict(inbound_stats, orient='index', columns = [ 'inbound'])
        outbound = pd.DataFrame.from_dict(outbound_stats, orient='index', columns = [ 'outbound'])

        return pd.concat([inbound.astype(int), outbound.astype(int)], axis=1, sort=False)
        
    def _calc_load_time(self, mov, unit='coy units', load_type='load'):
        "Function to calculate loading/offloading time according to the loading/offloading speed for each movement"
        if load_type == 'load':
            return (mov.loc['finished load time'] - pd.to_timedelta(
                mov.loc[unit]/(self.PPH_LOAD[process.extractOne(mov.loc['load type'], list(
                    self.PPH_LOAD.keys()))[0]]/60), unit='minute'))
        elif load_type == 'offload':
            return (mov.loc['offload time'] + pd.to_timedelta(
                mov.loc[unit]/(self.PPH_OFFLOAD[process.extractOne(mov.loc['load type'], list(
                    self.PPH_OFFLOAD.keys()))[0]]/60), unit='minute'))

    def inbound_unit_coy(self, break_point=None, num_years=0):
        "Convert inbound conveyables into units, e.g., bags for smalls & docs and calculate finished offloading time"
        cinbound_df = self.combine_inbound_df(break_point, num_years)
        cinbound_df.loc[:, 'coy units'] = cinbound_df['volume'] * (cinbound_df['box'] + (
            cinbound_df['documents'] + cinbound_df['smalls'])/self.PIECE_PER_BAG)
        inbound_coy_df = cinbound_df.loc[cinbound_df['coy units']!=0]
        inbound_coy_df = inbound_coy_df.assign(temp = inbound_coy_df.apply(
            self._calc_load_time, axis=1, unit='coy units', load_type='offload').dt.round('1s'))
        inbound_coy_df.rename(columns={'temp': 'finished offload time'}, inplace=True)

        return inbound_coy_df

    def inbound_unit_ncy(self, break_point=None, num_years=0):
        "Convert inbound nonconveyables into units, e.g., cages and calculate finished offloading time"
        cinbound_df = self.combine_inbound_df(break_point, num_years)
        cinbound_df.loc[:, 'ncy units'] = cinbound_df['volume'] * (cinbound_df['dangerous goods']/self.PARCEL_PER_CAGE_DG  \
            + cinbound_df['ncoy']/self.PARCEL_PER_CAGE_NCY + cinbound_df['hpc']/self.HPC + cinbound_df['freight'])
        inbound_ncy_df = cinbound_df.loc[cinbound_df['ncy units']!=0]
        inbound_ncy_df = inbound_ncy_df.assign(temp = inbound_ncy_df.apply(
            self._calc_load_time, axis=1, unit='ncy units', load_type='offload').dt.round('1s'))
        inbound_ncy_df.rename(columns={'temp': 'finished offload time'}, inplace=True)

        return inbound_ncy_df

    def outbound_unit_coy(self, break_point=None, num_years=0):
        "Convert outbound conveyables into units, e.g., bags for smalls & docs and calculate loading time"
        coutbound_df = self.combine_outbound_df(break_point, num_years)
        coutbound_df.loc[:, 'coy units'] = coutbound_df['volume'] * (coutbound_df['box'] + (coutbound_df['documents'] \
            + coutbound_df['smalls'])/self.PIECE_PER_BAG)
        outbound_coy_df = coutbound_df.loc[coutbound_df['coy units']!=0]
        outbound_coy_df = outbound_coy_df.assign(temp = outbound_coy_df.apply(
            self._calc_load_time, axis=1, unit='coy units', load_type='load').dt.round('1s'))
        outbound_coy_df.rename(columns={'temp': 'load time'}, inplace=True)

        return outbound_coy_df

    def outbound_unit_ncy(self, break_point=None, num_years=0):
        "Convert outbound nonconveyables into units, e.g., cages and calculate loading time"
        coutbound_df = self.combine_outbound_df(break_point, num_years)
        coutbound_df.loc[:, 'ncy units'] = coutbound_df['volume'] * (coutbound_df['dangerous goods']/self.PARCEL_PER_CAGE_DG \
            + coutbound_df['ncoy']/self.PARCEL_PER_CAGE_NCY + coutbound_df['hpc']/self.HPC + coutbound_df['freight'])
        outbound_ncy_df = coutbound_df.loc[coutbound_df['ncy units']!=0]
        outbound_ncy_df = outbound_ncy_df.assign(temp = outbound_ncy_df.apply(
            self._calc_load_time, axis=1, unit='ncy units', load_type='load').dt.round('1s'))
        outbound_ncy_df.rename(columns={'temp': 'load time'}, inplace=True)

        return outbound_ncy_df

    def _volume_spread(self, movement, ts='15T'):
        "Function to calculate the offloading time windows"
        start_time = movement.loc['offload time'].floor(ts)
        end_time = movement.loc['finished offload time'].floor(ts)
        if ts[-1] == 'T':
            k = int(ts[:-1])/60
        elif ts[-1] == 'H':
            k = int(ts[:-1])
        else:
            raise ValueError("Invalid format for time span")
        pph = self.PPH_OFFLOAD[process.extractOne(movement.loc['load type'], list(self.PPH_OFFLOAD.keys()))[0]]
        speed = pph * k
        unit_var = process.extractOne('unit', movement.index)[0]
        for time in pd.date_range(start=start_time, end=end_time, freq=ts):
            if start_time == end_time:
                movement.loc[time] = min(movement.loc[unit_var], speed)
                for n in range(1, int(movement.loc[unit_var]/speed) + 2):
                    leftover = movement.loc[unit_var] - n*speed
                    movement.loc[time + pd.Timedelta(str(int(ts[:-1])*n) + ts[-1])] = max(0, leftover)
            else:
                if time == start_time:
                    movement.loc[time] = (time + pd.Timedelta(ts) - movement.loc['offload time']).seconds/60 * (pph/60)
                elif time == end_time:
                    movement.loc[time] = (movement.loc['finished offload time'] - time).seconds/60 * (pph/60)
                else: 
                    movement.loc[time] = speed
        return movement

    def volume_availability_coy(self, break_point=None, ts='15T', num_years=0):
        "Calculate volume availability for all the conveyables"
        volume_coy = self.inbound_unit_coy(break_point, num_years).loc[:, [
            'origin', 'arrival time', 'offload time', 'finished offload time', 'coy units', 'load type']]
        start_time = (min(volume_coy['offload time'])).floor(ts)
        end_time = (max(volume_coy['finished offload time'])).ceil(ts)

        # generate the whole availability array for every 15 mins
        t_index = pd.date_range(start=start_time, end=end_time, freq=ts)
        for col in list(t_index): volume_coy.loc[:, col] = 0

        return volume_coy.apply(self._volume_spread, ts=ts, axis=1, result_type='broadcast').applymap(
            lambda x: round(x) if type(x)==float else x)  # standardize the number of digits in the very beginning

    def volume_availability_sd(self, break_point=None, ts='15T', num_years=0):
        "Calculate volume availability for all the smalls and documents for their sort capacity simulation later"
        volume_sd = self.inbound_unit_coy(break_point, num_years).loc[:, ['finished offload time', 'volume', 'smalls', 'documents']]
        volume_sd.loc[:, 'sd'] = (volume_sd['smalls'] + volume_sd['documents'])*volume_sd['volume']
        volume_sd = volume_sd.loc[volume_sd['sd']!=0]
        volume_sd.set_index('finished offload time', inplace=True)
        volume_sd.drop(columns=['smalls', 'documents', 'volume'], inplace=True)
        volume_sd.sort_index(inplace=True)
        volume_sd = volume_sd.groupby(['finished offload time']).agg({'sd': 'sum'})
        volume_sta = volume_sd.resample(ts).sum().round()
        return volume_sta

    def volume_availability_ncy(self, break_point=None, ts='15T', num_years=0):
        "Calculate volume availability for all the non-conveyables"
        volume_ncy = self.inbound_unit_ncy(break_point, num_years).loc[:, [
            'offload time', 'finished offload time', 'ncy units', 'load type']]
        start_time = (min(volume_ncy['offload time'])).floor(ts)
        end_time = (max(volume_ncy['finished offload time'])).ceil(ts)

        # generate the whole availability array for every 15 mins
        t_index = pd.date_range(start=start_time, end=end_time, freq=ts)
        for col in list(t_index): volume_ncy.loc[:, col] = 0

        return volume_ncy.apply(self._volume_spread, ts=ts, axis=1, result_type='broadcast').applymap(
            lambda x: round(x) if type(x)==float else x) 

    def inbound_doors_coy(self, break_point=None, ts='15T', num_years=0):
        "Calculate inbound doors for all the conveyables"
        volume_availability = self.volume_availability_coy(break_point, ts, num_years)
        time_index = [t for t in volume_availability.columns if type(t) is pd.Timestamp]
        volume_spread = volume_availability[['load type'] + time_index].set_index('load type')

        inbound_doors = dict()
        for load_type in volume_spread.index.unique():
            volume_coy = volume_spread.loc[[load_type]]
            doors = max(volume_coy.apply(np.count_nonzero, axis=0))
            inbound_doors[load_type] = doors
        
        return inbound_doors

    def dest_distribution_outbound_coy(self, break_point=None, station=False, num_years=0):
        "Calculate destination distribution purely based on the outbound data for conveyables"
        dest_df = self.outbound_unit_coy(break_point, num_years).loc[:, ['destination', 'departure time', 'load type', 'volume']]
        dest_df = dest_df.applymap(lambda x: x.lower().strip() if type(x) == str else x)
        dest_df.reset_index(inplace=True)
        dest_df.loc[:, 'unique destination'] = dest_df['level_0']  + '-' + dest_df['load type'] + '-' + dest_df['destination']
        inbound_coy = self.inbound_unit_coy(break_point, num_years).loc[:, ['origin', 'arrival time']]
        inbound_coy.reset_index(inplace=True)
        inbound_coy.loc[:, 'id'] = inbound_coy['origin'] + '-' + inbound_coy['arrival time'].map(lambda x: x.strftime('%H%M'))
        inbound_coy.set_index('id', inplace=True)
        min_dep_time = pd.DataFrame()
        distribution_df = pd.DataFrame()
        max_dep_time = max(dest_df['departure time'])

        for mov in inbound_coy.itertuples(index=True, name=None):
            dest = dest_df.copy()
            dest.loc[dest['departure time'] < mov[-1], 'volume'] = 0
            dest.loc[dest['departure time'] < mov[-1], 'departure time'] = max_dep_time
            if station:
                dest.loc[dest['level_0'] == mov[1], 'volume'] = 0
                dest.loc[dest['level_0'] == mov[1], 'departure time'] = max_dep_time
            distribution = dest.groupby('unique destination').agg({'volume': sum, 'departure time': min})
            distribution.loc[:, 'percent'] = distribution.loc[:, 'volume']/distribution.loc[:, 'volume'].sum()
            vol_dist = pd.DataFrame(distribution['percent']).T
            vol_dist.index = (mov[0], )
            distribution_df = distribution_df.append(vol_dist, sort=True)
            dep_time = pd.DataFrame(distribution['departure time']).T
            dep_time.index = (mov[0], )
            min_dep_time = min_dep_time.append(dep_time, sort=True)
        
        min_dep_time = min_dep_time.min(axis=1).reset_index()
        min_dep_time.columns = ['inbound key', 'departure time']

        return min_dep_time, distribution_df

    def vol_dist_coy(self, break_point=None, num_years=0):
        "Calculate volume distribution based on movements if provided, more accurate than the orig-dest relationship"
        outbound_coy = self.outbound_unit_coy(break_point, num_years).loc[:, ['destination', 'departure time', 'load type']]
        inbound_coy = self.inbound_unit_coy(break_point, num_years).loc[:, ['origin', 'arrival time']]
        vol_dist = self.get_vol_dist_df()
        outbound_coy.loc[:, 'departure time'] = outbound_coy.loc[:, 'departure time'].map(lambda x: x.strftime('%H%M'))
        inbound_coy.loc[:, 'arrival time'] = inbound_coy.loc[:, 'arrival time'].map(lambda x: x.strftime('%H%M'))

        vol_dist_coy = pd.merge(inbound_coy, vol_dist, left_on=['origin', 'arrival time'], right_on=['From', 'ArrTime'], how='left', sort=True)
        # populate relationships for inbound volumes movements: the same as their service movement peers
        vol_dist_coy = pd.concat([vol_dist_coy.dropna(), pd.merge(
            vol_dist_coy.loc[vol_dist_coy.isnull().any(axis=1), ['origin', 'arrival time']], vol_dist, left_on=[
                'origin'], right_on=['From'], how='left', sort=True)], ignore_index=True, sort=False)

        vol_dist_coy = pd.merge(
            vol_dist_coy, outbound_coy, left_on=['LocationTo', 'DepTime'], right_on=['destination', 'departure time'], how='left', sort=True)
        # populate relationships for outbound volumes movements: the same as their service movement peers
        vol_dist_coy = pd.concat([vol_dist_coy.dropna(), pd.merge(
            vol_dist_coy.loc[vol_dist_coy.isnull().any(axis=1), list(set(vol_dist_coy.columns) - set(outbound_coy.columns))], outbound_coy, left_on=[
                'LocationTo'], right_on=['destination'], how='left', sort=True)], ignore_index=True, sort=False)
        
        vol_dist_coy.loc[:, 'unique destination'] = vol_dist_coy['destination'] + '-' + vol_dist_coy['load type']

        vol_dist_coy.loc[:, 'inbound key'] = vol_dist_coy['origin'] + '-' + vol_dist_coy['arrival time']

        # remove all non-found results, not sure whether it is the correct way, need to reconsider and improve
        vol_dist_coy.dropna(inplace=True)

        if break_point is not None:
            break_point = int(break_point[-5:].replace(':', ''))
            vol_dist_coy.loc[:, 'departure time'] = vol_dist_coy.loc[:, 'departure time'].astype(int)
            vol_dist_coy.loc[vol_dist_coy['departure time'] < break_point, 'departure time'] += 2400

        min_dep_time = vol_dist_coy[['inbound key', 'departure time']].groupby('inbound key').agg({'departure time': min})

        distribution_df = pd.crosstab(index=vol_dist_coy['inbound key'], columns=vol_dist_coy['unique destination'], values=vol_dist_coy['Pieces'],\
            aggfunc=np.sum).fillna(0)

        return min_dep_time, distribution_df.apply(lambda x: x/x.sum(), axis=1)

    def sort_capacity_coy(self, sort_window: tuple, min_capacity=None, max_capacity=None, break_point=None, ts='15T', num_years=0):
        "Estimate sort capacity for all the conveyalables going through an automate sorter given the sort window (start and stop sorting time)"
        volume_availability = self.volume_availability_coy(break_point, ts, num_years)
        time_index = [t for t in volume_availability.columns if type(t) is pd.Timestamp]
        volume_spread = volume_availability[time_index].sum(axis=0).T

        start_time = pd.to_datetime(sort_window[0])
        end_time = pd.to_datetime(sort_window[1])
        # calculate the volume needed to be sorted per 15 mins within the sort window
        volume_2bsort = volume_spread[start_time : end_time]
        if volume_spread.index[0] > start_time:
            volume_2bsort[0] = volume_spread[start_time: volume_spread.index[0]].sum()
        else:
            volume_2bsort[0] = volume_spread[volume_spread.index[0]:start_time].sum()

        return solve_lp_model(volume_2bsort, self.MAX_BUFFER, min_capacity, max_capacity)

    def sort_capacity_sd(self, sort_window: tuple, min_capacity=None, max_capacity=None, break_point=None, ts='15T', num_years=0):
        "Estimate sort capacity for all the smalls&docs going through an automate sorter given the sort window (start and stop sorting time)"
        volume_spread = self.volume_availability_sd(break_point, ts, num_years)

        start_time = pd.to_datetime(sort_window[0])
        end_time = pd.to_datetime(sort_window[1])
        # calculate the volume needed to be sorted per 15 mins within the sort window
        volume_2bsort = volume_spread[start_time : end_time].squeeze()
        if volume_spread.index[0] > start_time:
            volume_2bsort[0] = volume_spread[start_time: volume_spread.index[0]].sum()
        else:
            volume_2bsort[0] = volume_spread[volume_spread.index[0]:start_time].sum()

        return solve_lp_model(volume_2bsort, self.MAX_BUFFER, min_capacity, max_capacity)

    def dest_spread_coy(self, sort_window, min_capacity=None, max_capacity=None, method='fifo', break_point=None, station=False, ts='15T', num_years=0):
        '''Calculate dataframe of destination spread for conveyables based on the origin-destination relationship and different methods, 
        There are five methods, depending on which volume distribution we have:
            random: randomly select from available movements to sort
            fifo: sort the movements that arrive earlier first, i.e., first in first sort
            earliest_dep_first: sort the movements that have the earliest departure time, only availabe when we have detailed data per movement, e.g., from Ali
        TO_DO: add service_first method - sort service movements first
        '''
        volume_sorted = self.sort_capacity_coy(sort_window, min_capacity, max_capacity, break_point, ts, num_years).loc[:, 'sorted']
        volume_sorted = volume_sorted[volume_sorted != 0]
        volume_availability = self.volume_availability_coy(break_point, ts, num_years)
        volume_availability.loc[:, 'id'] = volume_availability['origin'] + '-' + volume_availability['arrival time'].map(
            lambda x: x.strftime('%H%M')) # set unique id for each movement
        volume_availability.reset_index(inplace=True)
        if self.vol_dist_sheet:
            min_dep_time, dest_distribution = self.vol_dist_coy()
            volume_availability = pd.merge(
                volume_availability, min_dep_time, left_on=['id'], right_on=['inbound key'], how='left', sort=False)
            spread_func = lambda mov: (mov['sumtotal'] * dest_distribution.loc[mov['id'], :].T)
        elif self.orig_dest_sheet:
            dest_distribution = self.get_orig_dest_df()
            dest_distribution.set_index('origin', inplace=True)
            spread_func = lambda mov: (mov['sumtotal'] * dest_distribution.loc[mov['origin'], :].T)
        else:
            min_dep_time, dest_distribution = self.dest_distribution_outbound_coy(break_point, station, num_years)
            volume_availability = pd.merge(
                volume_availability, min_dep_time, left_on=['id'], right_on=['inbound key'], how='left', sort=False)
            spread_func = lambda mov: (mov['sumtotal'] * dest_distribution.loc[mov['id'], :].T)
        time_index = [t for t in volume_availability.columns if type(t) is pd.Timestamp]
        time_range = [time_index[0] - pd.Timedelta(minutes=15)] + list(volume_sorted.index)
        df_dest_spread = pd.DataFrame()
        df_inbound_dist = pd.DataFrame()
        xray_tput = {}
        for (index, value) in enumerate(time_range[:-1]):
            start_time, end_time = value, time_range[index + 1]
            volume_availability, dest_spread, inbound_dist, xray = calc_volume_spread(
                volume_sorted, volume_availability, dest_distribution, start_time, end_time, method, spread_func, ts)
            df_dest_spread = df_dest_spread.append(dest_spread, sort=True)
            df_inbound_dist = df_inbound_dist.append(inbound_dist, sort=True)
            xray_tput.update(xray)

        return df_dest_spread, df_inbound_dist, xray_tput

    def outbound_doors_coy(self, sort_window, min_capacity=None, max_capacity=None, method='fifo', break_point=None, station=False, ts='15T', num_years=0):
        "Function to calculate outbound doors for each destination"
        if ts[-1] == 'T':
            k = int(ts[:-1])/60
        elif ts[-1] == 'H':
            k = int(ts[:-1])
        else:
            raise ValueError("Invalid format for time span")

        df_dest_spread = self.dest_spread_coy(sort_window, min_capacity, max_capacity, method, break_point, station, ts, num_years)[0]

        df_dest_spread[df_dest_spread==0] = 1

        calc_door_dest = lambda dest: dest/(
            (self.PPH_LOAD[process.extractOne(dest.name.split('-')[1], list(self.PPH_LOAD.keys()))[0]]*k) + (
                self.OUTBOUND_BUFFER[process.extractOne(dest.name.split('-')[1], list(self.OUTBOUND_BUFFER.keys()))[0]]))

        dest_doors = df_dest_spread.apply(calc_door_dest, axis=0).applymap(lambda x: math.ceil(x))
        dest_doors.loc['max door per dest', :] = dest_doors.max(axis=0)
        dest_doors.loc[:, f'doors per {ts}'] = dest_doors.sum(axis=1)

        outbound_door = int(max(dest_doors.iloc[:-1, -1]))

        print(f'The number of outbound doors is {outbound_door}')

        return dest_doors.astype(int).loc[dest_doors.iloc[:-1, -1].idxmax()]

    def schedule_type_dist(self, sort_window, min_capacity=None, max_capacity=None, method='fifo', break_point=None, station=False, ts='15T', num_years=0):
        "Calculate the distributions of each sort for their inbound and outbound combinations"
        volume_sorted = self.sort_capacity_coy(sort_window, min_capacity, max_capacity, break_point, ts, num_years).loc[:, 'sorted']
        volume_sorted = pd.DataFrame(volume_sorted[volume_sorted != 0])

        df_dest_spread, inbound_distribution, xray_tput = self.dest_spread_coy(
            sort_window, min_capacity, max_capacity, method, break_point, station, ts, num_years)
        inbound_distribution = inbound_distribution.fillna(0).applymap(lambda x: '{0: .2f}%'.format(x * 100))
        inbound_distribution.columns = ['inbound-' + col for col in inbound_distribution.columns]
        xray_tput = pd.DataFrame.from_dict(xray_tput, orient='index', columns=[r'%Xray(linehaul to air)']).applymap(
            lambda x: '{0: .2f}%'.format(x * 100))

        for col in df_dest_spread.columns:
            replacement = {col: col.split('-')[0]}
            df_dest_spread.rename(columns = replacement, inplace=True)

        outbound_distribution = df_dest_spread.groupby(level=0, axis=1).sum()
        outbound_distribution = outbound_distribution.apply(lambda x: x/x.sum(), axis=1).applymap(lambda x: '{0: .2f}%'.format(x * 100))

        outbound_distribution.columns = ['outbound-' + col for col in outbound_distribution.columns]

        return volume_sorted.join([inbound_distribution, outbound_distribution, xray_tput], how='left')

    def export_to_excel(self, sort_window1, sort_window2=None, min_capacity=None, max_capacity=None, method='fifo', break_point=None, station=False, ts='15T', num_years=0, output_path=None):
        "Export the main results to an excel file"
        if output_path:
            output_path = Path(output_path)
        else:
            output_path = Path(__file__).parent

        wb = Workbook()
        # create the overview sheet
        ws = wb.active
        ws.title = "Overview"
        stats = self.descriptive_stats(break_point, num_years)
        for r in dataframe_to_rows(stats, index=True, header=True):
            ws.append(r)
        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        place_holder = 12
        for schedule_type in ['linehaul', 'pud', 'air', 'all']:
            if (self.schedule_sheets.get(schedule_type) is not None) or (schedule_type == 'all'):
                graph = self.hourly_volume_graph(schedule_type=schedule_type, break_point=break_point, num_years=num_years)
                volume_graph = graph.get_figure()
                volume_graph.savefig(output_path / Path(schedule_type + '_' + 'hourly_volume_graph.png'), dpi = 90)
                img = openpyxl.drawing.image.Image(output_path / Path(schedule_type + '_' + 'hourly_volume_graph.png'))
                img.anchor = 'A' + str(place_holder)
                place_holder += 25
                ws.add_image(img)

        # create the sort capacity coy sheet
        ws_sc = wb.create_sheet("Sort Capacity COY")
        sort_capacity_coy1 = self.sort_capacity_coy(sort_window1, min_capacity, max_capacity, break_point, ts, num_years)
        sorter_thput1 = int(math.ceil(sort_capacity_coy1['sorted'].max() / 500.0) * 500.0)
        sort_capacity_thput1 = self.sort_capacity_coy(sort_window1, sorter_thput1, sorter_thput1, break_point, ts, num_years)
        for r in dataframe_to_rows(sort_capacity_thput1, index=True, header=True):
            ws_sc.append(r)
        
        if sort_window2:
            sort_capacity_coy2 = self.sort_capacity_coy(sort_window2, min_capacity, max_capacity, break_point, ts, num_years)
            sorter_thput2 = int(math.ceil(sort_capacity_coy2['sorted'].max() / 500.0) * 500.0)
            sort_capacity_thput2 = self.sort_capacity_coy(sort_window2, sorter_thput2, sorter_thput2, break_point, ts, num_years)
            for r in dataframe_to_rows(sort_capacity_thput2, index=True, header=True):
                ws_sc.append(r)

        # create the inbound doors sheet
        ws_id = wb.create_sheet("Inbound Doors")
        ws_id['A1'] = 'Inbound Doors (Damir Method)'
        sorter_thput = max(sorter_thput1, sorter_thput2) if sort_window2 else sorter_thput1
        ws_id['B1'] = int((sorter_thput/self.PPH_OFFLOAD['LOOSE_TRAILER'])/0.72)
        ws_id['A4'] = 'Inbound Doors (Simultaneously Arrival)'
        inbound_doors_coy = self.inbound_doors_coy(break_point, ts, num_years)
        for next_row in range(5, len(inbound_doors_coy)+5):
            load_type, doors = inbound_doors_coy.popitem()
            ws_id.cell(column=1 , row=next_row, value=load_type)
            ws_id.cell(column=2 , row=next_row, value=doors)
        
        # create the outbound doors sheet
        ws_od = wb.create_sheet("Outbound Doors")
        ws_od['A1'] = 'Sort Window'
        ws_od['B1'] = sort_window1[0]
        ws_od['C1'] = sort_window1[1]
        outbound_doors_coy1 = self.outbound_doors_coy(sort_window1, sorter_thput1, sorter_thput1, method, break_point, station, ts, num_years)
        outbound_doors1 = outbound_doors_coy1.to_frame(name = 'Outbound Doors')
        for r in dataframe_to_rows(outbound_doors1, index=True, header=True):
            ws_od.append(r)
        
        if sort_window2:
            start_row = len(outbound_doors1) + 7
            ws_od['A' + str(start_row)] = 'Sort Window'
            ws_od['B' + str(start_row)] = sort_window2[0]
            ws_od['C' + str(start_row)] = sort_window2[0]
            outbound_doors_coy2 = self.outbound_doors_coy(sort_window2, sorter_thput2, sorter_thput2, method, break_point, station, ts, num_years)
            outbound_doors2 = outbound_doors_coy2.to_frame(name = 'Outbound Doors')
            for r in dataframe_to_rows(outbound_doors2, index=True, header=True):
                ws_od.append(r)

        # create the xray sheet
        if self.air_schedule_sheet:
            wx = wb.create_sheet("X-Ray Rate")
            dist1 = self.schedule_type_dist(sort_window1, sorter_thput1, sorter_thput1, method, break_point, station, ts, num_years)
            for r in dataframe_to_rows(dist1, index=True, header=True):
                wx.append(r)
            
            if sort_window2:
                dist2 = self.schedule_type_dist(sort_window2, sorter_thput2, sorter_thput2, method, break_point, station, ts, num_years)
                for r in dataframe_to_rows(dist2, index=True, header=True):
                    wx.append(r)

        file_name = output_path / Path('frm_results.xlsx')

        wb.save(file_name)

    def output(self):
        "Method to persist/show the calculation result."
        pass

    def execute(self):
        "Method to execute whole FRM task"
        self.output()