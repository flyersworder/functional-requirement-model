from frm import FRM
from openpyxl import load_workbook
from pathlib import Path
import pytest

@pytest.fixture
def frm_instance():
    file_path = Path(__file__)
    sample_excel = file_path.parents[0] / Path('sample_schedule_file.xlsx')
    print("Generate a FRM instance")
    yield FRM(sample_excel)
    print("Termination")

@pytest.fixture
def sample_lh_work_sheet():
    file_path = Path(__file__)
    sample_excel = file_path.parents[0] / Path('sample_schedule_file.xlsx')
    work_book = load_workbook(sample_excel, data_only=True)

    yield work_book['Input Road LH Schedule']
    print("Termination")

@pytest.fixture
def sample_pud_work_sheet():
    file_path = Path(__file__)
    sample_excel = file_path.parents[0] / Path('sample_schedule_file.xlsx')
    work_book = load_workbook(sample_excel, data_only=True)

    yield work_book['Input Road PUD schedule']
    print("Termination")